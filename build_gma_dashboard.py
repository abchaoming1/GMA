import json
import math
import os
import re
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
CURRENT_MONTH = datetime.now().month
FIRST_PAYMENT_RATE = 0.2375


def to_float(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if math.isnan(value) if isinstance(value, float) else False:
            return 0.0
        return float(value)
    text = str(value).replace("$", "").replace(",", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def normalize_2025_date(value):
    text = to_text(value)
    if text.startswith("2026-"):
        return "2025-" + text[5:]
    return text


def rows_from_sheet(workbook, sheet_name):
    ws = workbook[sheet_name]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    return rows[2:]


def parse_price(value):
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "")
    numbers = re.findall(r"\d+(?:\.\d+)?", text)
    if not numbers:
        return 0.0
    return float(numbers[-1])


def parse_2025(rows):
    parsed = []
    current_group = None
    group_id = 0
    for idx, row in enumerate(rows, start=1):
        if not any(v not in (None, "") for v in row):
            continue
        month = int(to_float(row[2])) if len(row) > 2 and to_float(row[2]) else 0
        sku = to_text(row[4]) if len(row) > 4 else ""
        if not sku:
            continue

        qty = to_float(row[9]) if len(row) > 9 else 0.0
        group_total_qty = to_float(row[10]) if len(row) > 10 else 0.0
        gross = to_float(row[11]) if len(row) > 11 else 0.0
        group_total_income = to_float(row[12]) if len(row) > 12 else 0.0
        group_first_payment = to_float(row[13]) if len(row) > 13 else 0.0
        group_returns_ratio = to_float(row[14]) if len(row) > 14 else 0.0
        group_returns_amount = to_float(row[15]) if len(row) > 15 else 0.0
        group_second_payment = to_float(row[16]) if len(row) > 16 else 0.0
        group_actual_rev = to_float(row[17]) if len(row) > 17 else 0.0
        is_group_start = bool(group_total_income or group_actual_rev or group_first_payment or group_second_payment)
        if is_group_start:
            group_id += 1
            current_group = {
                "id": f"2025-g{group_id}",
                "total_income": group_total_income or gross,
                "total_qty": group_total_qty or qty,
                "first_payment": group_first_payment,
                "returns_ratio": group_returns_ratio,
                "returns_amount": group_returns_amount,
                "second_payment": group_second_payment,
                "actual_rev": group_actual_rev,
                "month": month,
            }
        elif current_group:
            group_total_qty = current_group["total_qty"]
            group_total_income = current_group["total_income"]
            group_first_payment = current_group["first_payment"]
            group_returns_ratio = current_group["returns_ratio"]
            group_returns_amount = current_group["returns_amount"]
            group_second_payment = current_group["second_payment"]
            group_actual_rev = current_group["actual_rev"]
        allocation_weight = 1.0
        if current_group and current_group["total_income"]:
            allocation_weight = gross / current_group["total_income"]
        net_allocated = 0.0
        if current_group and current_group["actual_rev"] and current_group["total_income"]:
            net_allocated = current_group["actual_rev"] * allocation_weight
        elif group_actual_rev:
            net_allocated = group_actual_rev
        first_allocated = (current_group["first_payment"] * allocation_weight) if current_group else 0.0
        second_allocated = (current_group["second_payment"] * allocation_weight) if current_group else 0.0

        parsed.append(
            {
                "year": 2025,
                "groupId": current_group["id"] if current_group else f"2025-g{group_id}",
                "isGroupStart": is_group_start,
                "month": month,
                "type": to_text(row[1]) if len(row) > 1 else "",
                "date": normalize_2025_date(row[3]) if len(row) > 3 else "",
                "sku": sku,
                "model": to_text(row[5]) if len(row) > 5 else "",
                "priceText": to_text(row[6]) if len(row) > 6 else "",
                "promoPrice": to_float(row[7]) if len(row) > 7 else parse_price(row[6] if len(row) > 6 else ""),
                "inventory": to_float(row[8]) if len(row) > 8 else 0.0,
                "qty": qty,
                "totalQty": group_total_qty,
                "grossRev": gross,
                "totalIncome": gross,
                "groupTotalIncome": group_total_income,
                "originalFirstPayment": group_first_payment,
                "originalReturnsRatio": group_returns_ratio,
                "originalReturnsAmount": group_returns_amount,
                "originalSecondPayment": group_second_payment,
                "originalActualRev": group_actual_rev,
                "firstPayment": first_allocated,
                "secondPayment": second_allocated,
                "returnsRatio": current_group["returns_ratio"] if current_group else 0.0,
                "returnsAmount": current_group["returns_amount"] if current_group else 0.0,
                "actualRev": net_allocated,
                "netRev": net_allocated,
            }
        )
    return parsed


def parse_2026(rows):
    parsed = []
    current_group_rows = []
    groups = []

    def flush_group():
        nonlocal current_group_rows
        if current_group_rows:
            groups.append(current_group_rows)
            current_group_rows = []

    for idx, row in enumerate(rows, start=1):
        if not any(v not in (None, "") for v in row):
            continue
        month = int(to_float(row[2])) if len(row) > 2 and to_float(row[2]) else 0
        uses_new_layout = len(row) > 18 and bool(to_text(row[3])) and bool(to_text(row[8]))
        if uses_new_layout:
            status = to_text(row[3])
            start_date = to_text(row[4])
            end_date = to_text(row[5])
            sku = to_text(row[8])
            model = to_text(row[7])
            price = to_float(row[10])
            promo = parse_price(row[11])
            begin_inventory = to_float(row[12])
            end_inventory = to_float(row[13])
            qty = to_float(row[14])
            ach_rate = to_float(row[15])
            total_qty = to_float(row[16])
            group_total_income = to_float(row[17])
            group_actual_rev = to_float(row[18])
        else:
            status = ""
            start_date = to_text(row[3]) if len(row) > 3 else ""
            end_date = to_text(row[4]) if len(row) > 4 else ""
            sku = to_text(row[5]) if len(row) > 5 else ""
            model = to_text(row[6]) if len(row) > 6 else ""
            price = to_float(row[7]) if len(row) > 7 else 0.0
            promo = parse_price(row[8] if len(row) > 8 else "")
            begin_inventory = to_float(row[9]) if len(row) > 9 else 0.0
            end_inventory = to_float(row[10]) if len(row) > 10 else 0.0
            qty = to_float(row[11]) if len(row) > 11 else 0.0
            ach_rate = to_float(row[12]) if len(row) > 12 else 0.0
            total_qty = to_float(row[13]) if len(row) > 13 else 0.0
            group_total_income = to_float(row[14]) if len(row) > 14 else 0.0
            group_actual_rev = to_float(row[15]) if len(row) > 15 else 0.0
        if not sku:
            continue
        type_name = to_text(row[1]) if len(row) > 1 else ""

        starts_actual_group = bool(group_actual_rev or total_qty or group_total_income)
        if starts_actual_group:
            flush_group()

        row_total_income = qty * promo if qty and promo else 0.0

        item = {
            "id": f"r{idx}",
            "year": 2026,
            "month": month,
            "type": type_name,
            "status": status,
            "startDate": start_date,
            "endDate": end_date,
            "sku": sku,
            "model": model,
            "price": price,
            "promoPrice": promo,
            "beginInventory": begin_inventory,
            "endInventory": end_inventory,
            "qty": qty,
            "achRate": ach_rate,
            "groupQty": total_qty,
            "groupGross": group_total_income,
            "groupTotalIncome": group_total_income,
            "groupNet": group_actual_rev,
            "groupActualRev": group_actual_rev,
            "grossRev": row_total_income,
            "totalIncome": row_total_income,
            "firstPayment": 0.0,
            "secondPayment": 0.0,
            "actualRev": 0.0,
            "netRev": 0.0,
            "forecastQty": 0.0,
            "forecastTotalIncome": 0.0,
            "forecastFirstPayment": 0.0,
            "forecastSecondPayment": 0.0,
            "forecastActualRev": 0.0,
            "forecastNet": 0.0,
            "allocationWeight": 0.0,
            "isActual": False,
        }
        parsed.append(item)
        current_group_rows.append(item)
    flush_group()

    for group in groups:
        group_actual_rev = next((item["groupActualRev"] for item in group if item["groupActualRev"]), 0.0)
        group_total_income = next((item["groupTotalIncome"] for item in group if item["groupTotalIncome"]), 0.0)
        group_qty = next((item["groupQty"] for item in group if item["groupQty"]), 0.0)
        qty_sum = sum(item["qty"] for item in group)
        gross_sum = sum(item["qty"] * item["promoPrice"] for item in group)
        if not group_actual_rev and not group_total_income:
            continue
        group_first_payment = group_total_income * FIRST_PAYMENT_RATE if group_total_income else 0.0
        group_second_payment = max(group_total_income - group_first_payment - group_actual_rev, 0.0)
        for item in group:
            weight = 0.0
            if gross_sum:
                weight = (item["qty"] * item["promoPrice"]) / gross_sum
            elif group_qty:
                weight = item["qty"] / group_qty
            elif qty_sum:
                weight = item["qty"] / qty_sum
            item["totalIncome"] = (group_total_income * weight) if group_total_income else item["grossRev"]
            item["firstPayment"] = group_first_payment * weight
            item["secondPayment"] = group_second_payment * weight
            item["actualRev"] = group_actual_rev * weight
            item["netRev"] = item["actualRev"]
            item["isActual"] = item["netRev"] > 0
            if item["qty"] and item["promoPrice"]:
                item["grossRev"] = item["qty"] * item["promoPrice"]

    return parsed


def parse_inventory(rows):
    parsed = []
    for idx, row in enumerate(rows, start=1):
        if not any(v not in (None, "") for v in row):
            continue
        sku = to_text(row[4]) if len(row) > 4 else ""
        if not sku:
            continue
        parsed.append(
            {
                "month": int(to_float(row[2])) if len(row) > 2 and to_float(row[2]) else 0,
                "date": to_text(row[3]) if len(row) > 3 else "",
                "sku": sku,
                "model": to_text(row[5]) if len(row) > 5 else "",
                "originalInventory": to_float(row[6]) if len(row) > 6 else 0.0,
                "transferInventory": to_float(row[7]) if len(row) > 7 else 0.0,
                "totalInventory": to_float(row[8]) if len(row) > 8 else 0.0,
            }
        )
    return parsed


def sum_by_month(rows, field):
    result = {str(m): 0.0 for m in range(1, 13)}
    for row in rows:
        month = int(row.get("month") or 0)
        if 1 <= month <= 12:
            result[str(month)] += to_float(row.get(field))
    return result


def enrich_forecast(rows_2025, rows_2026, target_growth):
    baseline_by_month = {str(m): 0.0 for m in range(1, 13)}
    for row in rows_2025:
        month = int(row.get("month") or 0)
        if 1 <= month <= 12 and row.get("isGroupStart"):
            baseline_by_month[str(month)] += to_float(row.get("groupTotalIncome"))
    annual_baseline = sum(baseline_by_month.values())
    annual_target = annual_baseline * (1 + target_growth)
    actual_income = sum(row.get("totalIncome", 0.0) for row in rows_2026)
    remaining = max(annual_target - actual_income, 0.0)
    net_ratio = sum(row.get("netRev", 0.0) for row in rows_2025) / max(sum(row.get("totalIncome", 0.0) for row in rows_2025), 1.0)
    historical_second_payment_rate = sum(row.get("secondPayment", 0.0) for row in rows_2025) / max(
        sum(row.get("totalIncome", 0.0) for row in rows_2025), 1.0
    )
    actual_2026_income = sum(row.get("totalIncome", 0.0) for row in rows_2026 if row.get("isActual"))
    actual_2026_second = sum(row.get("secondPayment", 0.0) for row in rows_2026 if row.get("isActual"))
    second_payment_rate = (
        actual_2026_second / actual_2026_income if actual_2026_income else historical_second_payment_rate
    )
    profit_rate = max(1 - FIRST_PAYMENT_RATE - second_payment_rate, 0.01)
    known_prices = [row["promoPrice"] for row in rows_2026 if row.get("promoPrice")] + [
        row["promoPrice"] for row in rows_2025 if row.get("promoPrice")
    ]
    avg_promo_price = sum(known_prices) / len(known_prices) if known_prices else 80.0

    for row in rows_2026:
        row.setdefault("effectivePromoPrice", row.get("promoPrice") or row.get("price") or avg_promo_price)
        row["forecastQty"] = 0.0
        row["forecastTotalIncome"] = 0.0
        row["forecastFirstPayment"] = 0.0
        row["forecastSecondPayment"] = 0.0
        row["forecastActualRev"] = 0.0
        row["forecastNet"] = 0.0
        row["allocationWeight"] = 0.0

    return {
        "annualBaseline": annual_baseline,
        "annualTarget": annual_target,
        "actualIncome": actual_income,
        "remainingTarget": remaining,
        "netRatio": net_ratio,
        "firstPaymentRate": FIRST_PAYMENT_RATE,
        "secondPaymentRate": second_payment_rate,
        "profitRate": profit_rate,
        "avgPromoPrice": avg_promo_price,
        "baselineByMonth": baseline_by_month,
    }


def build_state(workbook_path):
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    rows_2025 = parse_2025(rows_from_sheet(wb, "2025-GMA"))
    rows_2026 = parse_2026(rows_from_sheet(wb, "2026-GMA"))
    inventory = parse_inventory(rows_from_sheet(wb, "2026-GMA-仍有库存"))
    target_growth = 0.5
    metrics = enrich_forecast(rows_2025, rows_2026, target_growth)

    monthly_expectations = {}
    average_monthly_target = metrics["annualTarget"] / 12
    for month in range(1, 13):
        actual = sum(row["totalIncome"] for row in rows_2026 if row["month"] == month)
        baseline = sum(row["groupTotalIncome"] for row in rows_2025 if row["month"] == month and row.get("isGroupStart"))
        monthly_expectations[str(month)] = {
            "month": month,
            "label": MONTH_LABELS[month - 1],
            "baselineIncome": baseline,
            "rawTarget": baseline * (1 + target_growth),
            "actualIncome": actual,
            "expectation": average_monthly_target,
            "manual": False,
        }

    return {
        "source": {
            "file": workbook_path.name,
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "currentMonth": CURRENT_MONTH,
        },
        "targetGrowth": target_growth,
        "rows2025": rows_2025,
        "rows2026": rows_2026,
        "inventory2026": inventory,
        "monthly": monthly_expectations,
        "skuTargets": {},
        "skuIncomeTargets": {},
        "metrics": metrics,
    }


def html_template(initial_state):
    state_json = json.dumps(initial_state, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GMA 2026 销售预测看板</title>
  <style>
    :root {{
      --bg: #f6f7f3;
      --panel: #ffffff;
      --ink: #22242a;
      --muted: #68707a;
      --line: #dfe4dc;
      --green: #2c7a57;
      --teal: #127f8c;
      --amber: #c57b1c;
      --rose: #b4475c;
      --blue: #3766a6;
      --soft-green: #e7f2ec;
      --soft-amber: #fff2da;
      --shadow: 0 10px 30px rgba(32, 38, 31, .08);
      font-family: Inter, "Segoe UI", "Microsoft YaHei", Arial, sans-serif;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--ink);
      background: var(--bg);
      font-size: 14px;
      letter-spacing: 0;
    }}
    header {{
      padding: 22px 28px 14px;
      border-bottom: 1px solid var(--line);
      background: #fbfcf8;
      position: sticky;
      top: 0;
      z-index: 10;
    }}
    .topbar {{
      display: flex;
      gap: 18px;
      align-items: end;
      justify-content: space-between;
      max-width: 1500px;
      margin: 0 auto;
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 24px;
      line-height: 1.2;
      font-weight: 780;
    }}
    .subtitle {{
      color: var(--muted);
      font-size: 13px;
    }}
    .controls {{
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      justify-content: flex-end;
      gap: 8px;
    }}
    button, .file-label {{
      border: 1px solid var(--line);
      background: var(--panel);
      color: var(--ink);
      min-height: 34px;
      padding: 7px 12px;
      border-radius: 6px;
      font-weight: 650;
      cursor: pointer;
      box-shadow: 0 1px 0 rgba(0,0,0,.02);
      white-space: nowrap;
    }}
    button.primary {{
      background: var(--green);
      color: white;
      border-color: var(--green);
    }}
    button:hover, .file-label:hover {{ border-color: #aab5aa; }}
    .hidden-file {{ display: none; }}
    main {{
      max-width: 1500px;
      margin: 0 auto;
      padding: 20px 28px 34px;
    }}
    .tabs {{
      display: flex;
      gap: 6px;
      border-bottom: 1px solid var(--line);
      margin-bottom: 18px;
      overflow-x: auto;
    }}
    .tab {{
      border: 0;
      border-radius: 0;
      box-shadow: none;
      background: transparent;
      color: var(--muted);
      padding: 11px 12px 10px;
      border-bottom: 3px solid transparent;
    }}
    .tab.active {{
      color: var(--green);
      border-bottom-color: var(--green);
    }}
    section.view {{ display: none; }}
    section.view.active {{ display: block; }}
    .grid {{
      display: grid;
      gap: 14px;
    }}
    .kpis {{
      grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }}
    .kpi {{
      padding: 14px 15px;
      min-height: 106px;
    }}
    .kpi-label {{
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 8px;
      white-space: nowrap;
    }}
    .kpi-value {{
      font-size: 23px;
      font-weight: 780;
      line-height: 1.15;
      margin-bottom: 8px;
    }}
    .kpi-sub {{
      color: var(--muted);
      font-size: 12px;
      line-height: 1.35;
    }}
    .dashboard-grid {{
      grid-template-columns: minmax(0, 1.5fr) minmax(320px, .8fr);
      margin-top: 14px;
    }}
    .panel-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 14px 16px 8px;
    }}
    h2 {{
      margin: 0;
      font-size: 16px;
      line-height: 1.25;
      font-weight: 760;
    }}
    .small-note {{
      color: var(--muted);
      font-size: 12px;
    }}
    .chart {{
      padding: 4px 12px 14px;
      min-height: 330px;
    }}
    .chart svg {{
      display: block;
      width: 100%;
      height: 320px;
    }}
    #paretoChart, #skuChart {{
      min-height: 405px;
    }}
    #paretoChart svg, #skuChart svg {{
      height: 390px;
    }}
    .progress-wrap {{
      padding: 4px 16px 18px;
    }}
    .progress-track {{
      height: 18px;
      background: #e9ede6;
      border-radius: 999px;
      overflow: hidden;
      border: 1px solid var(--line);
    }}
    .progress-bar {{
      height: 100%;
      width: 0;
      background: linear-gradient(90deg, var(--green), var(--teal));
      transition: width .25s ease;
    }}
    .summary-lines {{
      display: grid;
      gap: 8px;
      margin-top: 14px;
      color: var(--muted);
    }}
    .insight-list {{
      display: grid;
      gap: 8px;
      margin-top: 14px;
    }}
    .insight {{
      border-left: 3px solid var(--teal);
      background: #f7fbf8;
      padding: 8px 10px;
      color: #334047;
      line-height: 1.45;
      border-radius: 4px;
    }}
    .insight.risk {{
      border-left-color: var(--rose);
      background: #fff7f8;
    }}
    .conclusion-text {{
      padding: 2px 16px 16px;
      color: #27313a;
      font-size: 15px;
      line-height: 1.6;
      font-weight: 680;
    }}
    .line-item {{
      display: flex;
      justify-content: space-between;
      gap: 14px;
      border-bottom: 1px solid #edf0ea;
      padding-bottom: 7px;
    }}
    .line-item strong {{
      color: var(--ink);
      font-weight: 730;
      text-align: right;
    }}
    .table-wrap {{
      overflow: auto;
      border-top: 1px solid var(--line);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 900px;
    }}
    #dataTable {{
      min-width: 1760px;
    }}
    #data2025Table {{
      min-width: 1680px;
    }}
    #skuTable {{
      min-width: 1760px;
    }}
    #skuGapTable {{
      min-width: 1540px;
    }}
    th, td {{
      padding: 8px 9px;
      border-bottom: 1px solid #e8ece5;
      text-align: right;
      vertical-align: middle;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #f8faf5;
      color: #4d5660;
      font-size: 12px;
      font-weight: 720;
      z-index: 2;
    }}
    td.left, th.left {{ text-align: left; }}
    .num-input, .text-input, .select-input {{
      width: 100%;
      min-width: 72px;
      border: 1px solid transparent;
      background: transparent;
      border-radius: 5px;
      padding: 5px 6px;
      font: inherit;
      text-align: right;
      color: var(--ink);
    }}
    .text-input, .select-input {{ text-align: left; min-width: 120px; }}
    .num-input:focus, .text-input:focus, .select-input:focus {{
      outline: none;
      border-color: #8bb6a1;
      background: #fbfffb;
      box-shadow: 0 0 0 3px rgba(44, 122, 87, .12);
    }}
    .positive {{ color: var(--green); font-weight: 700; }}
    .negative {{ color: var(--rose); font-weight: 700; }}
    .risk-cell {{
      color: var(--rose);
      font-weight: 800;
      background: #fff4f6;
    }}
    .pill {{
      display: inline-flex;
      align-items: center;
      gap: 4px;
      padding: 3px 8px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 720;
      background: var(--soft-green);
      color: var(--green);
    }}
    .pill.plan {{
      background: var(--soft-amber);
      color: var(--amber);
    }}
    .form-row {{
      display: grid;
      grid-template-columns: 160px 160px 1fr;
      gap: 10px;
      align-items: center;
      padding: 12px 16px;
      border-top: 1px solid var(--line);
    }}
    .form-row label {{
      color: var(--muted);
      font-weight: 700;
    }}
    .form-row input {{
      min-height: 34px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 6px 8px;
      text-align: right;
      font: inherit;
      background: #fff;
    }}
    .filter-row {{
      display: grid;
      grid-template-columns: minmax(220px, 1fr) 190px 160px minmax(180px, .8fr);
      gap: 10px;
      align-items: center;
      padding: 12px 16px;
    }}
    .filter-row input, .filter-row select {{
      min-height: 34px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 6px 8px;
      font: inherit;
      background: #fff;
      color: var(--ink);
    }}
    .check-control {{
      display: inline-flex;
      align-items: center;
      gap: 7px;
      color: var(--muted);
      font-weight: 700;
      white-space: nowrap;
    }}
    .check-control input {{
      width: 16px;
      height: 16px;
      accent-color: var(--green);
    }}
    .filter-summary {{
      text-align: right;
      color: var(--muted);
      font-size: 12px;
    }}
    .two-col {{
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
    }}
    .sku-grid {{
      grid-template-columns: minmax(0, .95fr) minmax(0, 1.05fr);
    }}
    .bars {{
      padding: 6px 16px 16px;
      display: grid;
      gap: 10px;
    }}
    .timeline-stats {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      padding: 2px 16px 12px;
    }}
    .stat-chip {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border: 1px solid var(--line);
      background: #fbfcf8;
      color: #334047;
      border-radius: 999px;
      padding: 5px 10px;
      font-size: 12px;
      font-weight: 720;
    }}
    .activity-timeline {{
      position: relative;
      display: grid;
      grid-template-columns: repeat(12, minmax(118px, 1fr));
      gap: 10px;
      padding: 8px 16px 16px;
      overflow-x: auto;
    }}
    .activity-timeline::before {{
      content: "";
      position: absolute;
      left: 16px;
      right: 16px;
      top: 43px;
      height: 2px;
      background: #dde7dd;
    }}
    .timeline-month {{
      position: relative;
      min-height: 148px;
      border: 1px solid var(--line);
      background: #fcfdfa;
      border-radius: 8px;
      padding: 28px 10px 10px;
    }}
    .timeline-month.has-activity {{
      background: #f7fbf8;
      border-color: #c7dacd;
    }}
    .timeline-dot {{
      position: absolute;
      top: -7px;
      left: 14px;
      width: 14px;
      height: 14px;
      border-radius: 50%;
      border: 3px solid #fff;
      background: #aeb9a9;
      box-shadow: 0 0 0 1px var(--line);
    }}
    .timeline-month.has-activity .timeline-dot {{
      background: var(--teal);
      box-shadow: 0 0 0 1px #9fc9c8;
    }}
    .timeline-top {{
      display: flex;
      align-items: start;
      justify-content: space-between;
      gap: 8px;
      margin-bottom: 8px;
    }}
    .timeline-label {{
      color: #334047;
      font-weight: 780;
    }}
    .timeline-count {{
      min-width: 34px;
      text-align: center;
      border-radius: 999px;
      background: var(--soft-green);
      color: var(--green);
      padding: 2px 7px;
      font-size: 12px;
      font-weight: 800;
    }}
    .timeline-count.empty {{
      background: #eef1eb;
      color: var(--muted);
    }}
    .timeline-chips {{
      display: flex;
      flex-wrap: wrap;
      gap: 5px;
      min-height: 22px;
      margin-bottom: 8px;
    }}
    .type-chip {{
      border-radius: 999px;
      background: #e8f1f2;
      color: #22666e;
      padding: 2px 7px;
      font-size: 11px;
      font-weight: 720;
    }}
    .timeline-events {{
      display: grid;
      gap: 5px;
      color: var(--muted);
      font-size: 11px;
      line-height: 1.35;
    }}
    .event-line {{
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: 150px 1fr 92px;
      gap: 10px;
      align-items: center;
      font-size: 12px;
    }}
    .bar-label {{
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
      font-weight: 650;
    }}
    .bar-track {{
      height: 14px;
      background: #e9ede6;
      border-radius: 999px;
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      background: var(--teal);
      border-radius: 999px;
    }}
    .status-dot {{
      display: inline-block;
      width: 8px;
      height: 8px;
      border-radius: 50%;
      margin-right: 6px;
      background: var(--green);
    }}
    .status-dot.plan {{ background: var(--amber); }}
    .empty-state {{
      padding: 24px;
      color: var(--muted);
    }}
    @media (max-width: 1100px) {{
      .kpis, .dashboard-grid, .two-col, .sku-grid {{ grid-template-columns: 1fr; }}
      .topbar {{ align-items: start; flex-direction: column; }}
      .controls {{ justify-content: flex-start; }}
    }}
    @media (max-width: 720px) {{
      header, main {{ padding-left: 14px; padding-right: 14px; }}
      .form-row {{ grid-template-columns: 1fr; }}
      .filter-row {{ grid-template-columns: 1fr; }}
      .filter-summary {{ text-align: left; }}
      .activity-timeline {{ grid-template-columns: repeat(12, minmax(118px, 132px)); }}
      .kpi-value {{ font-size: 20px; }}
      table {{ min-width: 760px; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="topbar">
      <div>
        <h1>GMA 2026 销售预测看板</h1>
        <div class="subtitle" id="sourceInfo"></div>
      </div>
      <div class="controls">
        <button class="primary" id="saveBtn">保存</button>
        <button id="allocateBtn">重置月度目标</button>
        <button id="exportBtn">导出Excel</button>
        <label class="file-label" for="importFile">导入数据</label>
        <input id="importFile" class="hidden-file" type="file" accept="application/json">
        <button id="resetBtn">恢复初始</button>
      </div>
    </div>
  </header>
  <main>
    <nav class="tabs" aria-label="Views">
      <button class="tab active" data-view="dashboard">总览</button>
      <button class="tab" data-view="monthly">月度目标</button>
      <button class="tab" data-view="sku">SKU 分析</button>
      <button class="tab" data-view="data2025">2025 数据</button>
      <button class="tab" data-view="data">2026 数据</button>
      <button class="tab" data-view="inventory">库存</button>
    </nav>

    <section id="dashboard" class="view active">
      <div class="grid kpis">
        <div class="panel kpi"><div class="kpi-label">2025 Total Income</div><div class="kpi-value" id="kpiBaselineIncome"></div><div class="kpi-sub">营收基线</div></div>
        <div class="panel kpi"><div class="kpi-label">2026 营收目标</div><div class="kpi-value" id="kpiTarget"></div><div class="kpi-sub"><span id="growthText"></span> 同比增长目标</div></div>
        <div class="panel kpi"><div class="kpi-label">2026 已录营收</div><div class="kpi-value" id="kpiIncome"></div><div class="kpi-sub" id="incomeSub">Total Income</div></div>
        <div class="panel kpi"><div class="kpi-label">营收缺口</div><div class="kpi-value" id="kpiGap"></div><div class="kpi-sub" id="gapSub"></div></div>
        <div class="panel kpi"><div class="kpi-label">2026 已录 QTY</div><div class="kpi-value" id="kpiQty"></div><div class="kpi-sub">销量口径</div></div>
        <div class="panel kpi"><div class="kpi-label">销量第一 SKU</div><div class="kpi-value" id="kpiTopSku"></div><div class="kpi-sub" id="topSkuSub"></div></div>
      </div>

      <div class="panel" style="margin-top:14px;">
        <div class="panel-head"><h2>本月结论</h2><span class="small-note">按当前月份与已录数据自动生成</span></div>
        <div id="monthlyConclusion" class="conclusion-text"></div>
      </div>

      <div class="grid dashboard-grid">
        <div class="panel">
          <div class="panel-head">
            <h2>月度营收目标 vs 实际</h2>
            <span class="small-note">2025 营收基线 / 2026 月度目标 / 2026 已录营收</span>
          </div>
          <div id="monthlyChart" class="chart"></div>
        </div>
        <div class="panel">
          <div class="panel-head"><h2>自动洞察</h2><span class="small-note" id="progressLabel"></span></div>
          <div class="progress-wrap">
            <div class="progress-track"><div class="progress-bar" id="progressBar"></div></div>
            <div class="summary-lines" id="summaryLines"></div>
            <div id="insightList" class="insight-list"></div>
          </div>
        </div>
      </div>

      <div class="grid two-col" style="margin-top:14px;">
        <div class="panel">
          <div class="panel-head"><h2>营收瀑布图</h2><span class="small-note">2025 基线、增长目标、2026 已录与缺口</span></div>
          <div id="waterfallChart" class="chart"></div>
        </div>
        <div class="panel">
          <div class="panel-head"><h2>SKU 销量 Pareto</h2><span class="small-note">按 2026 QTY 排名与累计占比</span></div>
          <div id="paretoChart" class="chart"></div>
        </div>
      </div>

      <div class="panel" style="margin-top:14px;">
        <div class="panel-head"><h2>月度活动时间轴</h2><span class="small-note">按活动类型 + 开始日期 + 结束日期去重统计</span></div>
        <div id="activityTimelineStats" class="timeline-stats"></div>
        <div id="activityTimeline" class="activity-timeline"></div>
      </div>

      <div class="grid two-col" style="margin-top:14px;">
        <div class="panel">
          <div class="panel-head"><h2>活动类型对比</h2><span class="small-note">按 Total Income 与 QTY 对比</span></div>
          <div id="typeChart" class="chart"></div>
        </div>
        <div class="panel">
          <div class="panel-head"><h2>SKU 销量 Top 10</h2><span class="small-note">按 2026 QTY 排名</span></div>
          <div id="skuBars" class="bars"></div>
        </div>
      </div>
    </section>

    <section id="monthly" class="view">
      <div class="panel">
        <div class="panel-head">
          <h2>以终为始目标拆解</h2>
          <span class="small-note">年度目标 = 2025 Total Income 营收 x (1 + 增长率)</span>
        </div>
        <div class="form-row">
          <label for="growthInput">增长率</label>
          <input id="growthInput" type="number" step="1" min="-100" max="500">
          <span class="small-note">当前目标为 2025 年 Total Income 营收增长 50%，月度目标可直接改。</span>
        </div>
        <div class="table-wrap">
          <table id="monthlyTable"></table>
        </div>
      </div>
    </section>

    <section id="sku" class="view">
      <div class="panel" style="margin-bottom:14px;">
        <div class="filter-row">
          <input id="skuSearchInput" type="search" placeholder="搜索 SKU / Model">
          <select id="skuSortSelect">
            <option value="incomeGap">按营收缺口</option>
            <option value="qtyGap">按 QTY 缺口</option>
            <option value="qty26">按 2026 QTY</option>
            <option value="income26">按 2026 Total Income</option>
            <option value="incomeCompletion">按营收完成率</option>
            <option value="qtyCompletion">按 QTY 完成率</option>
          </select>
          <label class="check-control"><input id="skuOnlyGapInput" type="checkbox">只看未达双目标</label>
          <div id="skuFilterSummary" class="filter-summary"></div>
        </div>
      </div>
      <div class="grid sku-grid">
        <div class="panel">
          <div class="panel-head"><h2>SKU 销量 Pareto</h2><span class="small-note">2026 销量排名与累计占比</span></div>
          <div id="skuChart" class="chart"></div>
        </div>
        <div class="panel">
          <div class="panel-head"><h2>SKU 明细</h2><span class="small-note">数量、收入、差距</span></div>
          <div class="table-wrap">
            <table id="skuTable"></table>
          </div>
        </div>
      </div>
      <div class="panel" style="margin-top:14px;">
        <div class="panel-head"><h2>未达目标 SKU</h2><span class="small-note">目标 = 2025 QTY / Total Income x (1 + 增长率)，两个目标都可单独改</span></div>
        <div class="table-wrap">
          <table id="skuGapTable"></table>
        </div>
      </div>
    </section>

    <section id="data2025" class="view">
      <div class="panel">
        <div class="panel-head">
          <h2>2025 活动与 SKU 数据</h2>
          <span class="small-note">完整展示 2025 原表口径，Rev 与 Total Income 分开列</span>
        </div>
        <div class="table-wrap">
          <table id="data2025Table"></table>
        </div>
      </div>
    </section>

    <section id="data" class="view">
      <div class="panel">
        <div class="panel-head">
          <h2>2026 活动与 SKU 数据</h2>
          <span class="small-note">数字改动后会即时重算；保存后留在本浏览器</span>
        </div>
        <div class="table-wrap">
          <table id="dataTable"></table>
        </div>
      </div>
    </section>

    <section id="inventory" class="view">
      <div class="panel">
        <div class="panel-head"><h2>2026 仍有库存</h2><span class="small-note">来自原 workbook 库存页</span></div>
        <div class="table-wrap">
          <table id="inventoryTable"></table>
        </div>
      </div>
    </section>
  </main>

  <script>
    const INITIAL_STATE = {state_json};
    const STORAGE_KEY = "gma-2026-dashboard-v8";
    let state = loadState();
    let activeView = "dashboard";
    const skuFilters = {{ search: "", sort: "incomeGap", onlyGap: false }};

    function clone(value) {{
      return JSON.parse(JSON.stringify(value));
    }}

    function normalizeState(nextState) {{
      nextState.monthly = nextState.monthly || {{}};
      nextState.skuTargets = nextState.skuTargets || {{}};
      nextState.skuIncomeTargets = nextState.skuIncomeTargets || {{}};
      nextState.rows2025 = nextState.rows2025 || [];
      nextState.rows2026 = nextState.rows2026 || [];
      nextState.inventory2026 = nextState.inventory2026 || [];
      nextState.source = nextState.source || clone(INITIAL_STATE.source || {{}});
      return nextState;
    }}

    function loadState() {{
      try {{
        const saved = localStorage.getItem(STORAGE_KEY);
        if (saved) return normalizeState(JSON.parse(saved));
      }} catch (error) {{}}
      return normalizeState(clone(INITIAL_STATE));
    }}

    function saveState() {{
      localStorage.setItem(STORAGE_KEY, JSON.stringify(state));
      flash("已保存");
    }}

    function flash(text) {{
      const btn = document.getElementById("saveBtn");
      const old = btn.textContent;
      btn.textContent = text;
      window.setTimeout(() => btn.textContent = old, 850);
    }}

    function money(value) {{
      const number = Number(value || 0);
      return "$" + number.toLocaleString("en-US", {{ maximumFractionDigits: 0 }});
    }}

    function num(value, digits = 0) {{
      return Number(value || 0).toLocaleString("en-US", {{ maximumFractionDigits: digits }});
    }}

    function pct(value, digits = 1) {{
      return (Number(value || 0) * 100).toFixed(digits) + "%";
    }}

    function byMonth(rows, field, filter = () => true) {{
      const result = Array.from({{ length: 13 }}, () => 0);
      rows.forEach(row => {{
        const month = Number(row.month || 0);
        if (month >= 1 && month <= 12 && filter(row)) result[month] += Number(row[field] || 0);
      }});
      return result;
    }}

    function monthly2025Income() {{
      const result = Array.from({{ length: 13 }}, () => 0);
      state.rows2025.forEach(row => {{
        const month = Number(row.month || 0);
        if (month >= 1 && month <= 12 && row.isGroupStart) result[month] += Number(row.groupTotalIncome || 0);
      }});
      return result;
    }}

    function baselineTotal() {{
      return state.rows2025.reduce((sum, row) => sum + Number(row.netRev || 0), 0);
    }}

    function baselineIncomeTotal() {{
      const originalTotal = state.rows2025
        .filter(row => row.isGroupStart)
        .reduce((sum, row) => sum + Number(row.groupTotalIncome || 0), 0);
      return originalTotal || state.rows2025.reduce((sum, row) => sum + Number(row.totalIncome || row.grossRev || 0), 0);
    }}

    function rawTargetTotal() {{
      return baselineIncomeTotal() * (1 + Number(state.targetGrowth || 0));
    }}

    function targetTotal() {{
      const monthly = Array.from({{ length: 12 }}, (_, index) => monthlyExpectation(index + 1)).reduce((sum, value) => sum + value, 0);
      return monthly || rawTargetTotal();
    }}

    function actualTotal() {{
      return actualIncomeTotal();
    }}

    function actualQtyTotal() {{
      return state.rows2026.reduce((sum, row) => sum + Number(row.qty || 0), 0);
    }}

    function actualIncomeTotal() {{
      return state.rows2026.reduce((sum, row) => sum + Number(row.totalIncome || row.grossRev || 0), 0);
    }}

    function actualPaymentTotal() {{
      return state.rows2026.reduce((sum, row) => sum + Number(row.firstPayment || 0) + Number(row.secondPayment || 0), 0);
    }}

    function actualProfitTotal() {{
      return state.rows2026.reduce((sum, row) => sum + Number(row.actualRev || row.netRev || 0), 0);
    }}

    function forecastRowsTotal() {{
      return state.rows2026.reduce((sum, row) => sum + Number(row.forecastNet || 0), 0);
    }}

    function monthlyExpectation(month) {{
      const entry = state.monthly[String(month)] || {{}};
      if (entry.manual) return Number(entry.expectation || 0);
      if (entry.expectation !== undefined) return Number(entry.expectation || 0);
      return rawTargetTotal() / 12;
    }}

    function defaultSkuTarget(row) {{
      return Number(row.qty25 || 0) * (1 + Number(state.targetGrowth || 0));
    }}

    function defaultSkuIncomeTarget(row) {{
      return Number(row.income25 || 0) * (1 + Number(state.targetGrowth || 0));
    }}

    function skuTargetQty(row) {{
      const saved = state.skuTargets ? state.skuTargets[row.sku] : undefined;
      if (saved === undefined || saved === null || saved === "") return defaultSkuTarget(row);
      return Number(saved || 0);
    }}

    function skuTargetIncome(row) {{
      const saved = state.skuIncomeTargets ? state.skuIncomeTargets[row.sku] : undefined;
      if (saved === undefined || saved === null || saved === "") return defaultSkuIncomeTarget(row);
      return Number(saved || 0);
    }}

    function setSkuTarget(sku, value) {{
      state.skuTargets = state.skuTargets || {{}};
      state.skuTargets[sku] = Math.max(Number(value || 0), 0);
    }}

    function setSkuIncomeTarget(sku, value) {{
      state.skuIncomeTargets = state.skuIncomeTargets || {{}};
      state.skuIncomeTargets[sku] = Math.max(Number(value || 0), 0);
    }}

    function projectedTotal() {{
      return actualIncomeTotal();
    }}

    function monthlyRows() {{
      const base = monthly2025Income();
      const actualByMonth = byMonth(state.rows2026, "totalIncome");
      const qtyActual = byMonth(state.rows2026, "qty");
      const incomeActual = byMonth(state.rows2026, "totalIncome");
      const firstActual = byMonth(state.rows2026, "firstPayment");
      const secondActual = byMonth(state.rows2026, "secondPayment");
      const growth = Number(state.targetGrowth || 0);
      return Array.from({{ length: 12 }}, (_, index) => {{
        const month = index + 1;
        const target = monthlyExpectation(month);
        return {{
          month,
          label: ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][index],
          baseline: base[month],
          rawTarget: base[month] * (1 + growth),
          actual: actualByMonth[month],
          qty: qtyActual[month],
          totalIncome: incomeActual[month],
          firstPayment: firstActual[month],
          secondPayment: secondActual[month],
          expectation: target,
          breakdown: target,
          gap: target - actualByMonth[month],
          manual: Boolean(state.monthly[String(month)] && state.monthly[String(month)].manual),
        }};
      }});
    }}

    function skuRows() {{
      const map = new Map();
      function get(sku, model) {{
        if (!map.has(sku)) map.set(sku, {{
          sku, model: model || "", qty25: 0, income25: 0, first25: 0, second25: 0, net25: 0,
          qty26: 0, income26: 0, first26: 0, second26: 0, actual26: 0, forecast26: 0, inventory: 0
        }});
        const item = map.get(sku);
        if (!item.model && model) item.model = model;
        return item;
      }}
      state.rows2025.forEach(row => {{
        const item = get(row.sku, row.model);
        item.qty25 += Number(row.qty || 0);
        item.income25 += Number(row.totalIncome || row.grossRev || 0);
        item.first25 += Number(row.firstPayment || 0);
        item.second25 += Number(row.secondPayment || 0);
        item.net25 += Number(row.netRev || 0);
      }});
      state.rows2026.forEach(row => {{
        const item = get(row.sku, row.model);
        item.qty26 += Number(row.qty || 0);
        item.income26 += Number(row.totalIncome || 0);
        item.first26 += Number(row.firstPayment || 0);
        item.second26 += Number(row.secondPayment || 0);
        item.actual26 += Number(row.netRev || 0);
        item.inventory += Number(row.beginInventory || 0);
      }});
      return Array.from(map.values()).map(item => {{
        item.net26 = item.actual26;
        item.qtyGrowth = item.qty25 ? item.qty26 / item.qty25 - 1 : null;
        item.incomeGrowth = item.income25 ? item.income26 / item.income25 - 1 : null;
        item.qtyTarget = skuTargetQty(item);
        item.qtyGap = Math.max(item.qtyTarget - item.qty26, 0);
        item.incomeTarget = skuTargetIncome(item);
        item.incomeGap = Math.max(item.incomeTarget - item.income26, 0);
        item.qtyCompletion = item.qty26 / Math.max(item.qtyTarget, 1);
        item.incomeCompletion = item.income26 / Math.max(item.incomeTarget, 1);
        item.gap = item.income26 - item.income25 * (1 + Number(state.targetGrowth || 0));
        return item;
      }}).sort((a, b) => b.qty26 - a.qty26);
    }}

    function underTargetSkuRows() {{
      return skuRows()
        .filter(row => (row.qtyTarget > 0 && row.qtyGap > 0) || (row.incomeTarget > 0 && row.incomeGap > 0))
        .sort((a, b) => (b.incomeGap - a.incomeGap) || (b.qtyGap - a.qtyGap));
    }}

    function sortedSkuRows(rows) {{
      const sortKey = skuFilters.sort || "incomeGap";
      const directions = {{
        incomeCompletion: 1,
        qtyCompletion: 1,
      }};
      return [...rows].sort((a, b) => {{
        if (sortKey === "sku") return String(a.sku || "").localeCompare(String(b.sku || ""));
        const direction = directions[sortKey] || -1;
        const av = Number(a[sortKey] || 0);
        const bv = Number(b[sortKey] || 0);
        if (av === bv) return (b.incomeGap - a.incomeGap) || (b.qtyGap - a.qtyGap) || String(a.sku || "").localeCompare(String(b.sku || ""));
        return direction * (av - bv);
      }});
    }}

    function matchesSkuFilter(row) {{
      const query = String(skuFilters.search || "").trim().toLowerCase();
      if (!query) return true;
      return `${{row.sku || ""}} ${{row.model || ""}}`.toLowerCase().includes(query);
    }}

    function visibleSkuRows() {{
      const rows = skuRows().filter(matchesSkuFilter);
      const scoped = skuFilters.onlyGap ? rows.filter(row => row.qtyGap > 0 || row.incomeGap > 0) : rows;
      return sortedSkuRows(scoped);
    }}

    function visibleUnderTargetSkuRows() {{
      return sortedSkuRows(underTargetSkuRows().filter(matchesSkuFilter));
    }}

    function renderSkuFilterSummary() {{
      const el = document.getElementById("skuFilterSummary");
      if (!el) return;
      const searchInput = document.getElementById("skuSearchInput");
      const sortSelect = document.getElementById("skuSortSelect");
      const onlyGapInput = document.getElementById("skuOnlyGapInput");
      if (searchInput && document.activeElement !== searchInput) searchInput.value = skuFilters.search;
      if (sortSelect) sortSelect.value = skuFilters.sort;
      if (onlyGapInput) onlyGapInput.checked = skuFilters.onlyGap;
      const all = skuRows();
      const matched = all.filter(matchesSkuFilter);
      const under = matched.filter(row => row.qtyGap > 0 || row.incomeGap > 0);
      el.textContent = `显示 ${{skuFilters.onlyGap ? under.length : matched.length}} / ${{all.length}} SKU，未达双目标 ${{under.length}}`;
    }}

    function averagePromoPrice() {{
      const prices = [
        ...state.rows2026.map(row => Number(row.promoPrice || 0)).filter(Boolean),
        ...state.rows2025.map(row => Number(row.promoPrice || 0)).filter(Boolean)
      ];
      return prices.length ? prices.reduce((sum, value) => sum + value, 0) / prices.length : 80;
    }}

    function typeRows() {{
      const map = new Map();
      state.rows2026.forEach(row => {{
        const type = row.type || "未填写";
        if (!map.has(type)) map.set(type, {{ type, income: 0, net: 0, qty: 0 }});
        const item = map.get(type);
        item.income += Number(row.totalIncome || 0);
        item.net += Number(row.netRev || 0);
        item.qty += Number(row.qty || 0);
      }});
      return Array.from(map.values()).filter(row => row.income > 0 || row.qty > 0).sort((a, b) => b.income - a.income);
    }}

    function shortDate(value) {{
      return String(value || "").replace(/^2026-/, "").replace(/^2025-/, "");
    }}

    function activityTimelineRows() {{
      const labels = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
      const months = Array.from({{ length: 12 }}, (_, index) => ({{
        month: index + 1,
        label: labels[index],
        activities: [],
        categoryMap: new Map(),
      }}));
      const events = new Map();
      state.rows2026.forEach(row => {{
        const month = Number(row.month || 0);
        if (month < 1 || month > 12) return;
        const type = row.type || "未填写";
        const startDate = row.startDate || "";
        const endDate = row.endDate || "";
        const status = row.status || "";
        const key = `${{month}}|${{type}}|${{startDate}}|${{endDate}}|${{status}}`;
        if (!events.has(key)) {{
          events.set(key, {{
            month,
            type,
            startDate,
            endDate,
            status,
            qty: 0,
            income: 0,
            skus: new Set(),
            models: new Set(),
          }});
        }}
        const event = events.get(key);
        event.qty += Number(row.qty || 0);
        event.income += Number(row.totalIncome || row.grossRev || 0);
        if (row.sku) event.skus.add(row.sku);
        if (row.model) event.models.add(row.model);
      }});
      Array.from(events.values()).forEach(event => {{
        event.skuList = Array.from(event.skus);
        event.modelList = Array.from(event.models);
        event.skuCount = event.skuList.length;
        delete event.skus;
        delete event.models;
        const month = months[event.month - 1];
        month.activities.push(event);
        month.categoryMap.set(event.type, (month.categoryMap.get(event.type) || 0) + 1);
      }});
      return months.map(month => {{
        const categories = Array.from(month.categoryMap.entries())
          .map(([type, count]) => ({{ type, count }}))
          .sort((a, b) => b.count - a.count || a.type.localeCompare(b.type));
        month.activities.sort((a, b) => String(a.startDate).localeCompare(String(b.startDate)) || a.type.localeCompare(b.type));
        return {{
          month: month.month,
          label: month.label,
          count: month.activities.length,
          categories,
          activities: month.activities,
        }};
      }});
    }}

    function currentMonthConclusion() {{
      const month = Number((state.source && state.source.currentMonth) || (new Date().getMonth() + 1));
      const row = monthlyRows().find(item => item.month === month);
      if (!row) return "暂无本月目标数据，请先确认月度目标表。";
      const completion = row.actual / Math.max(row.breakdown, 1);
      const status = completion >= 1 ? "已达成目标" : (completion >= .8 ? "接近目标" : "未达目标");
      const monthRows = state.rows2026.filter(item => Number(item.month || 0) === month);
      const skuMap = new Map();
      const typeMap = new Map();
      monthRows.forEach(item => {{
        const qty = Number(item.qty || 0);
        const income = Number(item.totalIncome || item.grossRev || 0);
        if (qty > 0 || income > 0) {{
          const skuKey = item.sku || "未填写 SKU";
          if (!skuMap.has(skuKey)) skuMap.set(skuKey, {{ sku: skuKey, qty: 0, income: 0 }});
          skuMap.get(skuKey).qty += qty;
          skuMap.get(skuKey).income += income;
          const typeKey = item.type || "未填写类型";
          if (!typeMap.has(typeKey)) typeMap.set(typeKey, {{ type: typeKey, qty: 0, income: 0 }});
          typeMap.get(typeKey).qty += qty;
          typeMap.get(typeKey).income += income;
        }}
      }});
      const topSku = Array.from(skuMap.values()).sort((a, b) => b.qty - a.qty)[0];
      const topType = Array.from(typeMap.values()).sort((a, b) => b.income - a.income)[0];
      const gapText = row.gap > 0 ? `还需补齐 ${{money(row.gap)}}` : `已超出目标 ${{money(Math.abs(row.gap))}}`;
      if (row.actual <= 0) {{
        return `${{month}}月尚未录入 Total Income，本月营收目标为 ${{money(row.breakdown)}}，建议先确认本月活动与 SKU 数据是否完整。`;
      }}
      return `${{month}}月当前 Total Income 为 ${{money(row.actual)}}，目标 ${{money(row.breakdown)}}，完成率 ${{pct(completion, 1)}}，状态为${{status}}；销量第一 SKU 是 ${{topSku ? topSku.sku : "-"}}（${{topSku ? num(topSku.qty) : 0}} units），主要活动类型是 ${{topType ? topType.type : "-"}}，${{gapText}}。`;
    }}

    function recalculate2025Allocations() {{
      const groups = new Map();
      state.rows2025.forEach(row => {{
        const key = row.groupId || `${{row.month}}-${{row.type}}-${{row.date}}`;
        if (!groups.has(key)) groups.set(key, []);
        groups.get(key).push(row);
      }});
      groups.forEach(group => {{
        const head = group[0] || {{}};
        const groupIncome = Number(head.groupTotalIncome || 0);
        const groupFirst = Number(head.originalFirstPayment || 0);
        const groupSecond = Number(head.originalSecondPayment || 0);
        const groupActual = Number(head.originalActualRev || 0);
        const grossSum = group.reduce((sum, row) => sum + Number(row.grossRev || 0), 0);
        group.forEach(row => {{
          const weight = groupIncome ? Number(row.grossRev || 0) / groupIncome : (grossSum ? Number(row.grossRev || 0) / grossSum : 0);
          row.totalIncome = Number(row.grossRev || 0);
          row.firstPayment = groupFirst * weight;
          row.secondPayment = groupSecond * weight;
          row.actualRev = groupActual * weight;
          row.netRev = row.actualRev;
        }});
      }});
    }}

    function renderAll() {{
      renderSource();
      renderKpis();
      renderMonthlyConclusion();
      renderMonthlyChart();
      renderProgress();
      renderWaterfallChart();
      renderParetoChart("paretoChart");
      renderActivityTimeline();
      renderTypeChart();
      renderQuantityBars("skuBars", skuRows().filter(r => r.qty26 > 0).slice(0, 10).map(r => [r.sku, r.qty26]));
      if (activeView === "monthly") renderMonthlyTable();
      if (activeView === "sku") {{
        renderSkuFilterSummary();
        renderParetoChart("skuChart");
        renderSkuTable();
        renderSkuGapTable();
      }}
      if (activeView === "data2025") renderData2025Table();
      if (activeView === "data") renderDataTable();
      if (activeView === "inventory") renderInventoryTable();
    }}

    function renderSource() {{
      document.getElementById("sourceInfo").textContent = `源文件：${{state.source.file}} ｜ 生成：${{state.source.generatedAt}}`;
    }}

    function renderKpis() {{
      const target = targetTotal();
      const actual = actualTotal();
      const gap = target - actual;
      const topSku = skuRows().find(row => row.qty26 > 0);
      document.getElementById("kpiBaselineIncome").textContent = money(baselineIncomeTotal());
      document.getElementById("kpiTarget").textContent = money(target);
      document.getElementById("kpiQty").textContent = num(actualQtyTotal());
      document.getElementById("kpiIncome").textContent = money(actualIncomeTotal());
      document.getElementById("kpiGap").textContent = money(Math.max(gap, 0));
      document.getElementById("kpiGap").className = "kpi-value " + (gap <= 0 ? "positive" : "negative");
      document.getElementById("incomeSub").textContent = `完成率 ${{pct(actual / Math.max(target, 1), 1)}}`;
      document.getElementById("gapSub").textContent = gap <= 0 ? "已达成营收目标" : "需要后续活动补齐";
      document.getElementById("kpiTopSku").textContent = topSku ? topSku.sku : "-";
      document.getElementById("topSkuSub").textContent = topSku ? `${{num(topSku.qty26)}} units` : "暂无销量";
      document.getElementById("growthText").textContent = pct(state.targetGrowth, 0);
      const growthInput = document.getElementById("growthInput");
      if (growthInput) growthInput.value = Math.round(Number(state.targetGrowth || 0) * 100);
    }}

    function renderMonthlyConclusion() {{
      const el = document.getElementById("monthlyConclusion");
      if (el) el.textContent = currentMonthConclusion();
    }}

    function svgEl(tag, attrs = {{}}, text = "") {{
      const el = document.createElementNS("http://www.w3.org/2000/svg", tag);
      Object.entries(attrs).forEach(([key, value]) => el.setAttribute(key, value));
      if (text !== "") el.textContent = text;
      return el;
    }}

    function renderMonthlyChart() {{
      const container = document.getElementById("monthlyChart");
      const rows = monthlyRows();
      const width = container.clientWidth || 900;
      const height = 320;
      const margin = {{ top: 18, right: 18, bottom: 44, left: 72 }};
      const innerW = width - margin.left - margin.right;
      const innerH = height - margin.top - margin.bottom;
      const max = Math.max(...rows.flatMap(r => [r.baseline, r.breakdown, r.actual]), 1);
      const svg = svgEl("svg", {{ viewBox: `0 0 ${{width}} ${{height}}`, role: "img" }});
      svg.appendChild(svgEl("rect", {{ x: 0, y: 0, width, height, fill: "transparent" }}));
      for (let i = 0; i <= 4; i += 1) {{
        const y = margin.top + innerH - innerH * i / 4;
        svg.appendChild(svgEl("line", {{ x1: margin.left, y1: y, x2: width - margin.right, y2: y, stroke: "#e5ebe1", "stroke-width": 1 }}));
        svg.appendChild(svgEl("text", {{ x: margin.left - 10, y: y + 4, "text-anchor": "end", fill: "#68707a", "font-size": 11 }}, money(max * i / 4)));
      }}
      const groupW = innerW / rows.length;
      const barW = Math.max(5, Math.min(16, groupW / 5));
      const colors = ["#aeb9a9", "#c57b1c", "#127f8c"];
      rows.forEach((row, idx) => {{
        const x0 = margin.left + idx * groupW + groupW / 2;
        [row.baseline, row.breakdown, row.actual].forEach((value, series) => {{
          const barH = innerH * value / max;
          const x = x0 + (series - 1) * (barW + 3) - barW / 2;
          const y = margin.top + innerH - barH;
          svg.appendChild(svgEl("rect", {{ x, y, width: barW, height: Math.max(0, barH), rx: 2, fill: colors[series] }}));
        }});
        svg.appendChild(svgEl("text", {{ x: x0, y: height - 18, "text-anchor": "middle", fill: "#59616a", "font-size": 11 }}, row.label));
      }});
      const legend = [["2025 营收", colors[0]], ["2026 目标", colors[1]], ["2026 已录", colors[2]]];
      legend.forEach((item, idx) => {{
        const x = margin.left + idx * 94;
        svg.appendChild(svgEl("rect", {{ x, y: height - 10, width: 10, height: 10, fill: item[1], rx: 2 }}));
        svg.appendChild(svgEl("text", {{ x: x + 15, y: height - 1, fill: "#59616a", "font-size": 11 }}, item[0]));
      }});
      container.replaceChildren(svg);
    }}

    function renderProgress() {{
      const target = targetTotal();
      const actual = actualTotal();
      const completion = Math.min(actual / Math.max(target, 1), 1.2);
      document.getElementById("progressBar").style.width = Math.min(completion * 100, 100) + "%";
      document.getElementById("progressLabel").textContent = pct(actual / Math.max(target, 1), 1);
      const remaining = Math.max(target - actual, 0);
      const activeMonths = monthlyRows().filter(r => r.month >= state.source.currentMonth && r.gap > 0);
      const avgNeed = remaining / Math.max(activeMonths.length, 1);
      const topSku = skuRows().find(row => row.qty26 > 0);
      const topType = typeRows()[0];
      const topSkuGap = underTargetSkuRows()[0];
      document.getElementById("summaryLines").innerHTML = [
        ["年度营收目标", money(target)],
        ["已录营收", money(actual)],
        ["剩余营收缺口", money(remaining)],
        ["剩余目标月均", money(avgNeed)],
        ["已录 QTY", num(actualQtyTotal())],
        ["已录 First + Second Payment", money(actualPaymentTotal())],
      ].map(([label, value]) => `<div class="line-item"><span>${{label}}</span><strong>${{value}}</strong></div>`).join("");
      const insights = [
        actual >= target
          ? `已录营收已覆盖全年目标，当前超出 ${{money(actual - target)}}。`
          : `全年营收目标还差 ${{money(remaining)}}，按剩余目标月份平均每月需贡献 ${{money(avgNeed)}}。`,
        topSku ? `销量最高 SKU 是 ${{topSku.sku}}，已录 ${{num(topSku.qty26)}} units，贡献营收 ${{money(topSku.income26)}}。` : "当前 2026 尚无已录销量。",
        topType ? `活动类型里 ${{topType.type}} 营收最高，贡献 ${{money(topType.income)}}，销量 ${{num(topType.qty)}} units。` : "活动类型暂无可用数据。",
        topSkuGap ? `SKU 双目标缺口最大的是 ${{topSkuGap.sku}}：QTY 缺口 ${{num(topSkuGap.qtyGap)}}，Total Income 缺口 ${{money(topSkuGap.incomeGap)}}。` : "当前 SKU 双目标均已达成或未设置目标。",
        `辅助利润指标：已录 Actual Rev 为 ${{money(actualProfitTotal())}}，营收利润率 ${{pct(actualProfitTotal() / Math.max(actualIncomeTotal(), 1), 1)}}。`,
        `当前目标口径为 Total Income；未发生活动不做自动预测，只在月度目标里保留营收拆解。`,
      ];
      document.getElementById("insightList").innerHTML = insights
        .map((text, index) => `<div class="insight ${{index === 0 && actual < target ? "risk" : ""}}">${{text}}</div>`)
        .join("");
    }}

    function renderActivityTimeline() {{
      const container = document.getElementById("activityTimeline");
      const statsEl = document.getElementById("activityTimelineStats");
      if (!container || !statsEl) return;
      const rows = activityTimelineRows();
      const total = rows.reduce((sum, row) => sum + row.count, 0);
      const activeMonths = rows.filter(row => row.count > 0).length;
      const categoryMap = new Map();
      rows.forEach(row => row.categories.forEach(category => {{
        categoryMap.set(category.type, (categoryMap.get(category.type) || 0) + category.count);
      }}));
      const categories = Array.from(categoryMap.entries())
        .map(([type, count]) => ({{ type, count }}))
        .sort((a, b) => b.count - a.count || a.type.localeCompare(b.type));
      statsEl.innerHTML = [
        `全年活动 ${{num(total)}} 次`,
        `活跃月份 ${{num(activeMonths)}} 个`,
        categories.length ? `活动品类 ${{categories.map(item => `${{item.type}} x${{item.count}}`).join(" / ")}}` : "活动品类 暂无",
      ].map(text => `<span class="stat-chip">${{escapeAttr(text)}}</span>`).join("");
      container.innerHTML = rows.map(row => {{
        const chips = row.categories.length
          ? row.categories.map(category => `<span class="type-chip">${{escapeAttr(category.type)}} x${{num(category.count)}}</span>`).join("")
          : `<span class="type-chip">无活动</span>`;
        const details = row.activities.slice(0, 3).map(activity => {{
          const dateText = activity.startDate || activity.endDate ? `${{shortDate(activity.startDate)}}-${{shortDate(activity.endDate)}}` : "日期未填";
          const skuText = activity.skuCount ? `${{num(activity.skuCount)}} SKU` : "SKU 未填";
          const detail = `${{dateText}}｜${{activity.type}}｜${{skuText}}`;
          return `<div class="event-line" title="${{escapeAttr(detail)}}">${{escapeAttr(detail)}}</div>`;
        }}).join("");
        const more = row.activities.length > 3 ? `<div class="event-line">+${{num(row.activities.length - 3)}} 更多活动</div>` : "";
        return `
          <div class="timeline-month ${{row.count ? "has-activity" : ""}}">
            <span class="timeline-dot"></span>
            <div class="timeline-top">
              <div class="timeline-label">${{row.label}}</div>
              <div class="timeline-count ${{row.count ? "" : "empty"}}">${{num(row.count)}}</div>
            </div>
            <div class="timeline-chips">${{chips}}</div>
            <div class="timeline-events">${{details || `<div class="event-line">暂无活动</div>`}}${{more}}</div>
          </div>`;
      }}).join("");
    }}

    function renderBars(id, rows, colorName) {{
      const max = Math.max(...rows.map(row => row[1]), 1);
      const color = colorName === "green" ? "var(--green)" : "var(--teal)";
      const html = rows.length ? rows.map(row => {{
        const width = Math.max(2, row[1] / max * 100);
        return `<div class="bar-row"><div class="bar-label" title="${{row[0]}}">${{row[0]}}</div><div class="bar-track"><div class="bar-fill" style="width:${{width}}%;background:${{color}}"></div></div><div>${{money(row[1])}}</div></div>`;
      }}).join("") : `<div class="empty-state">暂无数据</div>`;
      document.getElementById(id).innerHTML = html;
    }}

    function renderQuantityBars(id, rows) {{
      const max = Math.max(...rows.map(row => row[1]), 1);
      const html = rows.length ? rows.map(row => {{
        const width = Math.max(2, row[1] / max * 100);
        return `<div class="bar-row"><div class="bar-label" title="${{row[0]}}">${{row[0]}}</div><div class="bar-track"><div class="bar-fill" style="width:${{width}}%;background:var(--teal)"></div></div><div>${{num(row[1])}}</div></div>`;
      }}).join("") : `<div class="empty-state">暂无销量数据</div>`;
      document.getElementById(id).innerHTML = html;
    }}

    function renderWaterfallChart() {{
      const container = document.getElementById("waterfallChart");
      if (!container) return;
      const baseline = baselineIncomeTotal();
      const growth = rawTargetTotal() - baseline;
      const target = targetTotal();
      const actual = actualTotal();
      const gap = Math.max(target - actual, 0);
      const rows = [
        {{ label: "2025", value: baseline, color: "#aeb9a9" }},
        {{ label: "+50%", value: growth, color: "#c57b1c" }},
        {{ label: "目标", value: target, color: "#3766a6" }},
        {{ label: "已录", value: actual, color: "#127f8c" }},
        {{ label: "缺口", value: gap, color: "#b4475c" }},
      ];
      const width = container.clientWidth || 680;
      const height = 320;
      const margin = {{ top: 22, right: 18, bottom: 48, left: 72 }};
      const innerW = width - margin.left - margin.right;
      const innerH = height - margin.top - margin.bottom;
      const max = Math.max(...rows.map(r => r.value), 1);
      const svg = svgEl("svg", {{ viewBox: `0 0 ${{width}} ${{height}}` }});
      for (let i = 0; i <= 4; i += 1) {{
        const y = margin.top + innerH - innerH * i / 4;
        svg.appendChild(svgEl("line", {{ x1: margin.left, y1: y, x2: width - margin.right, y2: y, stroke: "#e5ebe1" }}));
        svg.appendChild(svgEl("text", {{ x: margin.left - 8, y: y + 4, "text-anchor": "end", fill: "#68707a", "font-size": 11 }}, money(max * i / 4)));
      }}
      const slot = innerW / rows.length;
      const barW = Math.min(54, slot * .56);
      rows.forEach((row, index) => {{
        const x = margin.left + index * slot + (slot - barW) / 2;
        const h = innerH * row.value / max;
        const y = margin.top + innerH - h;
        svg.appendChild(svgEl("rect", {{ x, y, width: barW, height: h, fill: row.color, rx: 3 }}));
        svg.appendChild(svgEl("text", {{ x: x + barW / 2, y: y - 6, "text-anchor": "middle", fill: "#334047", "font-size": 11 }}, money(row.value)));
        svg.appendChild(svgEl("text", {{ x: x + barW / 2, y: height - 18, "text-anchor": "middle", fill: "#59616a", "font-size": 12 }}, row.label));
      }});
      container.replaceChildren(svg);
    }}

    function renderParetoChart(id) {{
      const container = document.getElementById(id);
      if (!container) return;
      const sourceRows = id === "skuChart"
        ? skuRows().filter(row => matchesSkuFilter(row) && (!skuFilters.onlyGap || row.qtyGap > 0 || row.incomeGap > 0))
        : skuRows();
      const allRows = sourceRows.filter(row => row.qty26 > 0).sort((a, b) => b.qty26 - a.qty26);
      const rows = allRows.slice(0, 12);
      const totalQty = allRows.reduce((sum, row) => sum + row.qty26, 0) || 1;
      const width = container.clientWidth || 680;
      const height = 390;
      const denseLabels = rows.length > 7;
      const margin = {{ top: 40, right: 78, bottom: denseLabels ? 122 : 96, left: 64 }};
      const innerW = width - margin.left - margin.right;
      const innerH = height - margin.top - margin.bottom;
      const maxQty = Math.max(...rows.map(row => row.qty26), 1);
      const svg = svgEl("svg", {{ viewBox: `0 0 ${{width}} ${{height}}` }});
      const halo = {{ stroke: "#fff", "stroke-width": 4, "paint-order": "stroke", "stroke-linejoin": "round" }};
      for (let i = 0; i <= 4; i += 1) {{
        const y = margin.top + innerH - innerH * i / 4;
        svg.appendChild(svgEl("line", {{ x1: margin.left, y1: y, x2: width - margin.right, y2: y, stroke: "#e5ebe1" }}));
        svg.appendChild(svgEl("text", {{ x: margin.left - 8, y: y + 4, "text-anchor": "end", fill: "#68707a", "font-size": 10 }}, num(maxQty * i / 4)));
      }}
      [0.5, 0.8, 1].forEach(value => {{
        const y = margin.top + innerH - innerH * value;
        svg.appendChild(svgEl("text", {{ x: width - margin.right + 8, y: y + 4, fill: value === 0.8 ? "#c57b1c" : "#68707a", "font-size": 10 }}, pct(value, 0)));
        if (value === 0.8) svg.appendChild(svgEl("line", {{ x1: margin.left, y1: y, x2: width - margin.right, y2: y, stroke: "#c57b1c", "stroke-dasharray": "4 4", opacity: .55 }}));
      }});
      const slot = innerW / Math.max(rows.length, 1);
      const barW = Math.max(8, Math.min(28, slot * .55));
      let cumulative = 0;
      const points = [];
      rows.forEach((row, index) => {{
        const x = margin.left + index * slot + slot / 2;
        const h = innerH * row.qty26 / maxQty;
        const y = margin.top + innerH - h;
        const barLabelY = Math.max(margin.top - 10, y - 8);
        svg.appendChild(svgEl("rect", {{ x: x - barW / 2, y, width: barW, height: h, fill: "#127f8c", rx: 2 }}));
        svg.appendChild(svgEl("text", {{ x, y: barLabelY, "text-anchor": "middle", fill: "#334047", "font-size": 10, ...halo }}, num(row.qty26)));
        if (denseLabels) {{
          const labelY = height - 58;
          svg.appendChild(svgEl("text", {{ x, y: labelY, "text-anchor": "end", transform: `rotate(-35 ${{x}} ${{labelY}})`, fill: "#59616a", "font-size": 10, ...halo }}, row.sku));
        }} else {{
          const parts = row.sku.split("-");
          const firstLine = parts.length > 2 ? parts.slice(0, 2).join("-") : row.sku;
          const secondLine = parts.length > 2 ? parts.slice(2).join("-") : "";
          const label = svgEl("text", {{ x, y: height - 58, "text-anchor": "middle", fill: "#59616a", "font-size": 10, ...halo }});
          label.appendChild(svgEl("tspan", {{ x, dy: 0 }}, firstLine));
          if (secondLine) label.appendChild(svgEl("tspan", {{ x, dy: 13 }}, secondLine));
          svg.appendChild(label);
        }}
        cumulative += row.qty26;
        const cumulativePct = cumulative / totalQty;
        const py = margin.top + innerH - innerH * cumulative / totalQty;
        points.push([x, py, cumulativePct, barLabelY]);
      }});
      if (points.length) {{
        svg.appendChild(svgEl("polyline", {{ points: points.map(p => `${{p[0]}},${{p[1]}}`).join(" "), fill: "none", stroke: "#c57b1c", "stroke-width": 2 }}));
        points.forEach((point, index) => {{
          svg.appendChild(svgEl("circle", {{ cx: point[0], cy: point[1], r: 3, fill: "#c57b1c" }}));
          let labelY = point[1] - 18;
          if (labelY < margin.top + 10) labelY = point[1] + 20;
          if (Math.abs(labelY - point[3]) < 18) labelY = point[1] + 20;
          svg.appendChild(svgEl("text", {{ x: point[0], y: labelY, "text-anchor": "middle", fill: "#9a5b12", "font-size": 10, ...halo }}, pct(point[2], 0)));
        }});
      }}
      svg.appendChild(svgEl("text", {{ x: width / 2, y: height - 13, "text-anchor": "middle", fill: "#68707a", "font-size": 11, ...halo }}, "柱：QTY（柱顶数字）｜ 线：累计销量占比（百分比标签）"));
      container.replaceChildren(svg);
    }}

    function renderTypeChart() {{
      const container = document.getElementById("typeChart");
      if (!container) return;
      const rows = typeRows();
      const width = container.clientWidth || 680;
      const height = 320;
      const margin = {{ top: 22, right: 70, bottom: 44, left: 92 }};
      const innerW = width - margin.left - margin.right;
      const rowH = (height - margin.top - margin.bottom) / Math.max(rows.length, 1);
      const maxIncome = Math.max(...rows.map(row => row.income), 1);
      const maxQty = Math.max(...rows.map(row => row.qty), 1);
      const svg = svgEl("svg", {{ viewBox: `0 0 ${{width}} ${{height}}` }});
      rows.forEach((row, index) => {{
        const y = margin.top + index * rowH + rowH * .22;
        svg.appendChild(svgEl("text", {{ x: margin.left - 8, y: y + 12, "text-anchor": "end", fill: "#334047", "font-size": 12 }}, row.type));
        svg.appendChild(svgEl("rect", {{ x: margin.left, y, width: innerW * row.income / maxIncome, height: rowH * .22, fill: "#127f8c", rx: 2 }}));
        svg.appendChild(svgEl("rect", {{ x: margin.left, y: y + rowH * .32, width: innerW * row.qty / maxQty, height: rowH * .22, fill: "#c57b1c", rx: 2 }}));
        svg.appendChild(svgEl("text", {{ x: width - margin.right + 8, y: y + 12, fill: "#59616a", "font-size": 11 }}, money(row.income)));
      }});
      svg.appendChild(svgEl("text", {{ x: margin.left, y: height - 10, fill: "#68707a", "font-size": 11 }}, "绿色：Total Income ｜ 黄色：QTY 相对量级"));
      container.replaceChildren(svg);
    }}

    function renderMonthlyTable() {{
      const rows = monthlyRows();
      const table = document.getElementById("monthlyTable");
      table.innerHTML = `
        <thead><tr>
          <th class="left">月份</th><th>2025 Total Income</th><th>2026 营收目标</th><th>2026 已录营收</th><th>营收缺口</th><th>完成率</th><th>QTY</th><th>First Payment</th><th>Second Payment</th>
        </tr></thead>
        <tbody>
          ${{rows.map(row => `
            <tr>
              <td class="left">${{row.label}}</td>
              <td>${{money(row.baseline)}}</td>
              <td><input class="num-input" type="number" step="100" value="${{Math.round(row.expectation)}}" data-month="${{row.month}}" data-month-field="expectation"></td>
              <td>${{money(row.actual)}}</td>
              <td class="${{row.gap <= 0 ? "positive" : "negative"}}">${{money(Math.max(row.gap, 0))}}</td>
              <td class="${{row.actual / Math.max(row.breakdown, 1) < .8 ? "risk-cell" : "positive"}}">${{pct(row.actual / Math.max(row.breakdown, 1), 1)}}</td>
              <td>${{num(row.qty)}}</td>
              <td>${{money(row.firstPayment)}}</td>
              <td>${{money(row.secondPayment)}}</td>
            </tr>
          `).join("")}}
        </tbody>`;
      table.querySelectorAll("input[data-month-field]").forEach(input => {{
        input.addEventListener("change", () => {{
          const month = input.dataset.month;
          state.monthly[month] = state.monthly[month] || {{}};
          state.monthly[month].expectation = Number(input.value || 0);
          state.monthly[month].manual = true;
          renderAll();
        }});
      }});
    }}

    function renderSkuChart() {{
      const top = skuRows().slice(0, 12);
      const container = document.getElementById("skuChart");
      const width = container.clientWidth || 620;
      const height = 320;
      const margin = {{ top: 12, right: 18, bottom: 28, left: 132 }};
      const innerW = width - margin.left - margin.right;
      const rowH = (height - margin.top - margin.bottom) / Math.max(top.length, 1);
      const max = Math.max(...top.flatMap(r => [r.net25, r.net26]), 1);
      const svg = svgEl("svg", {{ viewBox: `0 0 ${{width}} ${{height}}` }});
      top.forEach((row, idx) => {{
        const y = margin.top + idx * rowH + rowH * .18;
        svg.appendChild(svgEl("text", {{ x: margin.left - 8, y: y + rowH * .44, "text-anchor": "end", fill: "#3d444c", "font-size": 11 }}, row.sku));
        const w25 = innerW * row.net25 / max;
        const w26 = innerW * row.net26 / max;
        svg.appendChild(svgEl("rect", {{ x: margin.left, y, width: w25, height: Math.max(5, rowH * .22), fill: "#aeb9a9", rx: 2 }}));
        svg.appendChild(svgEl("rect", {{ x: margin.left, y: y + rowH * .32, width: w26, height: Math.max(5, rowH * .22), fill: "#127f8c", rx: 2 }}));
      }});
      svg.appendChild(svgEl("text", {{ x: margin.left, y: height - 6, fill: "#68707a", "font-size": 11 }}, "灰色：2025 ｜ 绿色：2026 实际+预测"));
      container.replaceChildren(svg);
    }}

    function renderSkuTable() {{
      const rows = visibleSkuRows();
      const table = document.getElementById("skuTable");
      table.innerHTML = `
        <thead><tr>
          <th class="left">SKU</th><th class="left">Model</th><th>2026 QTY</th><th>2025 QTY</th><th>目标 QTY</th><th>QTY 缺口</th><th>QTY 完成率</th><th>2026 Total Income</th><th>2025 Total Income</th><th>目标 Total Income</th><th>营收缺口</th><th>营收完成率</th><th>销量同比</th><th>营收同比</th><th>计划库存</th>
        </tr></thead>
        <tbody>
          ${{rows.map(row => `
            <tr>
              <td class="left">${{row.sku}}</td>
              <td class="left">${{row.model}}</td>
              <td>${{num(row.qty26)}}</td>
              <td>${{num(row.qty25)}}</td>
              <td><input class="num-input" type="number" step="1" min="0" value="${{round0(row.qtyTarget)}}" data-sku-target="${{escapeAttr(row.sku)}}"></td>
              <td class="${{row.qtyGap > 0 ? "negative" : "positive"}}">${{num(row.qtyGap)}}</td>
              <td class="${{row.qtyCompletion < .8 ? "risk-cell" : "positive"}}">${{pct(row.qtyCompletion, 1)}}</td>
              <td>${{money(row.income26)}}</td>
              <td>${{money(row.income25)}}</td>
              <td><input class="num-input" type="number" step="100" min="0" value="${{round0(row.incomeTarget)}}" data-sku-income-target="${{escapeAttr(row.sku)}}"></td>
              <td class="${{row.incomeGap > 0 ? "negative" : "positive"}}">${{money(row.incomeGap)}}</td>
              <td class="${{row.incomeCompletion < .8 ? "risk-cell" : "positive"}}">${{pct(row.incomeCompletion, 1)}}</td>
              <td class="${{row.qtyGrowth === null || row.qtyGrowth >= 0 ? "positive" : "negative"}}">${{row.qtyGrowth === null ? "NEW" : pct(row.qtyGrowth, 1)}}</td>
              <td class="${{row.incomeGrowth === null || row.incomeGrowth >= 0 ? "positive" : "negative"}}">${{row.incomeGrowth === null ? "NEW" : pct(row.incomeGrowth, 1)}}</td>
              <td>${{num(row.inventory)}}</td>
            </tr>
          `).join("")}}
        </tbody>`;
      bindSkuTargetInputs(table);
    }}

    function renderSkuGapTable() {{
      const rows = visibleUnderTargetSkuRows();
      const table = document.getElementById("skuGapTable");
      table.innerHTML = `
        <thead><tr>
          <th class="left">SKU</th><th class="left">Model</th><th>2025 QTY</th><th>目标 QTY</th><th>2026 QTY</th><th>QTY 缺口</th><th>QTY 完成率</th><th>2025 Total Income</th><th>目标 Total Income</th><th>2026 Total Income</th><th>营收缺口</th><th>营收完成率</th>
        </tr></thead>
        <tbody>
          ${{rows.length ? rows.map(row => `
            <tr>
              <td class="left">${{row.sku}}</td>
              <td class="left">${{row.model}}</td>
              <td>${{num(row.qty25)}}</td>
              <td><input class="num-input" type="number" step="1" min="0" value="${{round0(row.qtyTarget)}}" data-sku-target="${{escapeAttr(row.sku)}}"></td>
              <td>${{num(row.qty26)}}</td>
              <td class="negative">${{num(row.qtyGap)}}</td>
              <td class="${{row.qtyCompletion < .8 ? "risk-cell" : "positive"}}">${{pct(row.qtyCompletion, 1)}}</td>
              <td>${{money(row.income25)}}</td>
              <td><input class="num-input" type="number" step="100" min="0" value="${{round0(row.incomeTarget)}}" data-sku-income-target="${{escapeAttr(row.sku)}}"></td>
              <td>${{money(row.income26)}}</td>
              <td class="negative">${{money(row.incomeGap)}}</td>
              <td class="${{row.incomeCompletion < .8 ? "risk-cell" : "positive"}}">${{pct(row.incomeCompletion, 1)}}</td>
            </tr>
          `).join("") : `<tr><td class="left" colspan="12">当前没有低于 SKU 双目标的项目。</td></tr>`}}
        </tbody>`;
      bindSkuTargetInputs(table);
    }}

    function bindSkuTargetInputs(table) {{
      table.querySelectorAll("input[data-sku-target]").forEach(input => {{
        input.addEventListener("change", () => {{
          setSkuTarget(input.dataset.skuTarget, input.value);
          renderAll();
        }});
      }});
      table.querySelectorAll("input[data-sku-income-target]").forEach(input => {{
        input.addEventListener("change", () => {{
          setSkuIncomeTarget(input.dataset.skuIncomeTarget, input.value);
          renderAll();
        }});
      }});
    }}

    function renderData2025Table() {{
      const table = document.getElementById("data2025Table");
      table.innerHTML = `
        <thead><tr>
          <th>年份</th><th>月</th><th class="left">类型</th><th class="left">日期</th><th class="left">SKU</th><th class="left">Model</th><th class="left">Price</th><th>Promo Price</th><th>Inventory</th><th>QTY</th><th>Total QTY</th><th>Rev</th><th>2025 Total Income</th><th>First Payment to GMA</th><th>Second Payment to GMA</th><th>Actual Rev</th><th>Returns Ratio</th><th>Returns Amount</th>
        </tr></thead>
        <tbody>
          ${{state.rows2025.map((row, index) => `
            <tr>
              <td><input class="num-input" type="number" step="1" value="${{row.year || 2025}}" data-2025-index="${{index}}" data-field="year"></td>
              <td><input class="num-input" type="number" step="1" value="${{row.month || ""}}" data-2025-index="${{index}}" data-field="month"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.type || "")}}" data-2025-index="${{index}}" data-field="type"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.date || "")}}" data-2025-index="${{index}}" data-field="date"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.sku || "")}}" data-2025-index="${{index}}" data-field="sku"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.model || "")}}" data-2025-index="${{index}}" data-field="model"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.priceText || "")}}" data-2025-index="${{index}}" data-field="priceText"></td>
              <td><input class="num-input" type="number" step="0.01" value="${{round2(row.promoPrice)}}" data-2025-index="${{index}}" data-field="promoPrice"></td>
              <td><input class="num-input" type="number" step="1" value="${{round0(row.inventory)}}" data-2025-index="${{index}}" data-field="inventory"></td>
              <td><input class="num-input" type="number" step="1" value="${{round0(row.qty)}}" data-2025-index="${{index}}" data-field="qty"></td>
              <td><input class="num-input" type="number" step="1" value="${{row.totalQty ? round0(row.totalQty) : ""}}" data-2025-index="${{index}}" data-field="totalQty"></td>
              <td><input class="num-input" type="number" step="100" value="${{round0(row.grossRev)}}" data-2025-index="${{index}}" data-field="grossRev"></td>
              <td><input class="num-input" type="number" step="100" value="${{row.groupTotalIncome ? round0(row.groupTotalIncome) : ""}}" data-2025-index="${{index}}" data-field="groupTotalIncome"></td>
              <td><input class="num-input" type="number" step="100" value="${{row.originalFirstPayment ? round0(row.originalFirstPayment) : ""}}" data-2025-index="${{index}}" data-field="originalFirstPayment"></td>
              <td><input class="num-input" type="number" step="100" value="${{row.originalSecondPayment ? round0(row.originalSecondPayment) : ""}}" data-2025-index="${{index}}" data-field="originalSecondPayment"></td>
              <td><input class="num-input" type="number" step="100" value="${{row.originalActualRev ? round0(row.originalActualRev) : ""}}" data-2025-index="${{index}}" data-field="originalActualRev"></td>
              <td><input class="num-input" type="number" step="0.0001" value="${{row.originalReturnsRatio ? Number(row.originalReturnsRatio || 0) : ""}}" data-2025-index="${{index}}" data-field="originalReturnsRatio"></td>
              <td><input class="num-input" type="number" step="1" value="${{row.originalReturnsAmount ? round0(row.originalReturnsAmount) : ""}}" data-2025-index="${{index}}" data-field="originalReturnsAmount"></td>
            </tr>
          `).join("")}}
        </tbody>`;
      table.querySelectorAll("input").forEach(input => {{
        input.addEventListener("change", () => {{
          const row = state.rows2025[Number(input.dataset["2025Index"])];
          const field = input.dataset.field;
          if (["type", "date", "sku", "model", "priceText"].includes(field)) row[field] = input.value;
          else row[field] = Number(input.value || 0);
          const groupFields = ["totalQty", "groupTotalIncome", "originalFirstPayment", "originalSecondPayment", "originalActualRev", "originalReturnsRatio", "originalReturnsAmount"];
          if (groupFields.includes(field)) {{
            state.rows2025
              .filter(item => item.groupId === row.groupId)
              .forEach(item => item[field] = row[field]);
          }}
          recalculate2025Allocations();
          renderAll();
        }});
      }});
    }}

    function renderDataTable() {{
      const table = document.getElementById("dataTable");
      table.innerHTML = `
        <thead><tr>
          <th>状态</th><th>月</th><th class="left">类型</th><th class="left">SKU</th><th class="left">Model</th><th>Promo</th><th>期初库存</th><th>期末库存</th><th>QTY</th><th>Total Income</th><th>First Payment</th><th>Second Payment</th><th>Actual Rev</th>
        </tr></thead>
        <tbody>
          ${{state.rows2026.map((row, index) => `
            <tr>
              <td class="left"><span class="pill ${{Number(row.netRev || 0) ? "" : "plan"}}"><span class="status-dot ${{Number(row.netRev || 0) ? "" : "plan"}}"></span>${{Number(row.netRev || 0) ? "实际" : "计划"}}</span></td>
              <td><input class="num-input" type="number" step="1" value="${{row.month || ""}}" data-index="${{index}}" data-field="month"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.type || "")}}" data-index="${{index}}" data-field="type"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.sku || "")}}" data-index="${{index}}" data-field="sku"></td>
              <td class="left"><input class="text-input" value="${{escapeAttr(row.model || "")}}" data-index="${{index}}" data-field="model"></td>
              <td><input class="num-input" type="number" step="0.01" value="${{round2(row.promoPrice)}}" data-index="${{index}}" data-field="promoPrice"></td>
              <td><input class="num-input" type="number" step="1" value="${{round0(row.beginInventory)}}" data-index="${{index}}" data-field="beginInventory"></td>
              <td><input class="num-input" type="number" step="1" value="${{round0(row.endInventory)}}" data-index="${{index}}" data-field="endInventory"></td>
              <td><input class="num-input" type="number" step="1" value="${{round0(row.qty)}}" data-index="${{index}}" data-field="qty"></td>
              <td><input class="num-input" type="number" step="100" value="${{round0(row.totalIncome || row.grossRev)}}" data-index="${{index}}" data-field="totalIncome"></td>
              <td><input class="num-input" type="number" step="100" value="${{round0(row.firstPayment)}}" data-index="${{index}}" data-field="firstPayment"></td>
              <td><input class="num-input" type="number" step="100" value="${{round0(row.secondPayment)}}" data-index="${{index}}" data-field="secondPayment"></td>
              <td><input class="num-input" type="number" step="100" value="${{round0(row.actualRev || row.netRev)}}" data-index="${{index}}" data-field="actualRev"></td>
            </tr>
          `).join("")}}
        </tbody>`;
      table.querySelectorAll("input").forEach(input => {{
        input.addEventListener("change", () => {{
          const row = state.rows2026[Number(input.dataset.index)];
          const field = input.dataset.field;
          if (["type", "sku", "model"].includes(field)) row[field] = input.value;
          else row[field] = Number(input.value || 0);
          if (field === "actualRev") row.netRev = row.actualRev;
          if (field === "totalIncome") row.grossRev = row.totalIncome;
          const month = String(row.month || "");
          if (state.monthly[month] && !state.monthly[month].manual) {{
            state.monthly[month].expectation = undefined;
          }}
          renderAll();
        }});
      }});
    }}

    function renderInventoryTable() {{
      const table = document.getElementById("inventoryTable");
      table.innerHTML = `
        <thead><tr><th>月</th><th class="left">日期</th><th class="left">SKU</th><th class="left">Model</th><th>原有计划库存</th><th>其他渠道划拨</th><th>总计库存</th></tr></thead>
        <tbody>
          ${{state.inventory2026.map(row => `
            <tr>
              <td>${{row.month || ""}}</td>
              <td class="left">${{row.date || ""}}</td>
              <td class="left">${{row.sku}}</td>
              <td class="left">${{row.model}}</td>
              <td>${{num(row.originalInventory)}}</td>
              <td>${{num(row.transferInventory)}}</td>
              <td>${{num(row.totalInventory)}}</td>
            </tr>
          `).join("")}}
        </tbody>`;
    }}

    function round0(value) {{ return Math.round(Number(value || 0)); }}
    function round2(value) {{ return Math.round(Number(value || 0) * 100) / 100; }}
    function escapeAttr(value) {{
      return String(value).replaceAll("&", "&amp;").replaceAll('"', "&quot;").replaceAll("<", "&lt;").replaceAll(">", "&gt;");
    }}

    function redistributeTargets() {{
      const averageTarget = rawTargetTotal() / 12;
      for (let month = 1; month <= 12; month += 1) {{
        state.monthly[String(month)] = state.monthly[String(month)] || {{}};
        state.monthly[String(month)].expectation = averageTarget;
        state.monthly[String(month)].manual = false;
      }}
      renderAll();
    }}

    function xmlEscape(value) {{
      return String(value ?? "").replaceAll("&", "&amp;").replaceAll("<", "&lt;").replaceAll(">", "&gt;").replaceAll('"', "&quot;");
    }}

    function excelCell(value) {{
      const isNumber = typeof value === "number" && Number.isFinite(value);
      const type = isNumber ? "Number" : "String";
      return `<Cell><Data ss:Type="${{type}}">${{xmlEscape(isNumber ? Math.round(value * 100) / 100 : value)}}</Data></Cell>`;
    }}

    function excelSheet(name, rows) {{
      return `<Worksheet ss:Name="${{xmlEscape(name).slice(0, 31)}}"><Table>${{rows.map(row => `<Row>${{row.map(excelCell).join("")}}</Row>`).join("")}}</Table></Worksheet>`;
    }}

    function exportExcel() {{
      const monthly = monthlyRows();
      const sku = skuRows();
      const gaps = underTargetSkuRows();
      const activityRows = activityTimelineRows();
      const sheets = [
        excelSheet("Summary", [
          ["Metric", "Value"],
          ["2025 Total Income", baselineIncomeTotal()],
          ["2026 Revenue Target", targetTotal()],
          ["2026 Actual Total Income", actualIncomeTotal()],
          ["Revenue Gap", Math.max(targetTotal() - actualIncomeTotal(), 0)],
          ["2026 QTY", actualQtyTotal()],
          ["2026 Actual Rev", actualProfitTotal()],
          ["Profit Margin", actualProfitTotal() / Math.max(actualIncomeTotal(), 1)],
          ["Current Month Conclusion", currentMonthConclusion()],
        ]),
        excelSheet("Monthly Targets", [
          ["Month", "2025 Total Income", "2026 Target", "2026 Actual Income", "Gap", "Completion", "QTY", "First Payment", "Second Payment"],
          ...monthly.map(row => [row.label, row.baseline, row.breakdown, row.actual, Math.max(row.gap, 0), row.actual / Math.max(row.breakdown, 1), row.qty, row.firstPayment, row.secondPayment]),
        ]),
        excelSheet("SKU Analysis", [
          ["SKU", "Model", "2026 QTY", "2025 QTY", "Target QTY", "Gap QTY", "QTY Completion", "2026 Total Income", "2025 Total Income", "Target Total Income", "Income Gap", "Income Completion", "QTY Growth", "Income Growth", "Inventory"],
          ...sku.map(row => [row.sku, row.model, row.qty26, row.qty25, row.qtyTarget, row.qtyGap, row.qtyCompletion, row.income26, row.income25, row.incomeTarget, row.incomeGap, row.incomeCompletion, row.qtyGrowth ?? "NEW", row.incomeGrowth ?? "NEW", row.inventory]),
        ]),
        excelSheet("Under Target SKU", [
          ["SKU", "Model", "2025 QTY", "Target QTY", "2026 QTY", "Gap QTY", "QTY Completion", "2025 Total Income", "Target Total Income", "2026 Total Income", "Income Gap", "Income Completion"],
          ...gaps.map(row => [row.sku, row.model, row.qty25, row.qtyTarget, row.qty26, row.qtyGap, row.qtyCompletion, row.income25, row.incomeTarget, row.income26, row.incomeGap, row.incomeCompletion]),
        ]),
        excelSheet("2025 Data", [
          ["Year", "Month", "Type", "Date", "SKU", "Model", "Price", "Promo Price", "Inventory", "QTY", "Total QTY", "Rev", "Total Income", "First Payment", "Second Payment", "Actual Rev", "Returns Ratio", "Returns Amount"],
          ...state.rows2025.map(row => [row.year, row.month, row.type, row.date, row.sku, row.model, row.priceText, row.promoPrice, row.inventory, row.qty, row.totalQty, row.grossRev, row.groupTotalIncome, row.originalFirstPayment, row.originalSecondPayment, row.originalActualRev, row.originalReturnsRatio, row.originalReturnsAmount]),
        ]),
        excelSheet("2026 Data", [
          ["Year", "Month", "Type", "SKU", "Model", "Promo", "Begin Inventory", "End Inventory", "QTY", "Total Income", "First Payment", "Second Payment", "Actual Rev"],
          ...state.rows2026.map(row => [row.year, row.month, row.type, row.sku, row.model, row.promoPrice, row.beginInventory, row.endInventory, row.qty, row.totalIncome || row.grossRev, row.firstPayment, row.secondPayment, row.actualRev || row.netRev]),
        ]),
        excelSheet("Activity Timeline", [
          ["Month", "Activity Count", "Activity Types", "Activity Details"],
          ...activityRows.map(row => [
            row.label,
            row.count,
            row.categories.map(category => `${{category.type}} x${{category.count}}`).join("; "),
            row.activities.map(activity => `${{activity.startDate || ""}}-${{activity.endDate || ""}} | ${{activity.type}} | ${{activity.skuCount}} SKU | QTY ${{Math.round(activity.qty)}} | Income ${{Math.round(activity.income)}}`).join("; "),
          ]),
        ]),
        excelSheet("Inventory", [
          ["Month", "Date", "SKU", "Model", "Original Inventory", "Transfer Inventory", "Total Inventory"],
          ...state.inventory2026.map(row => [row.month, row.date, row.sku, row.model, row.originalInventory, row.transferInventory, row.totalInventory]),
        ]),
      ];
      const xml = `<?xml version="1.0" encoding="UTF-8"?>
<?mso-application progid="Excel.Sheet"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:html="http://www.w3.org/TR/REC-html40">${{sheets.join("")}}</Workbook>`;
      const blob = new Blob([xml], {{ type: "application/vnd.ms-excel;charset=utf-8" }});
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = "GMA_2026_销售预测看板.xls";
      a.click();
      URL.revokeObjectURL(url);
    }}

    document.querySelectorAll(".tab").forEach(tab => {{
      tab.addEventListener("click", () => {{
        activeView = tab.dataset.view;
        document.querySelectorAll(".tab").forEach(item => item.classList.toggle("active", item === tab));
        document.querySelectorAll(".view").forEach(view => view.classList.toggle("active", view.id === activeView));
        renderAll();
      }});
    }});

    document.getElementById("saveBtn").addEventListener("click", saveState);
    document.getElementById("allocateBtn").addEventListener("click", redistributeTargets);
    document.getElementById("resetBtn").addEventListener("click", () => {{
      if (confirm("恢复到 Excel 初始数据？")) {{
        state = normalizeState(clone(INITIAL_STATE));
        localStorage.removeItem(STORAGE_KEY);
        renderAll();
      }}
    }});
    document.getElementById("exportBtn").addEventListener("click", exportExcel);
    document.getElementById("importFile").addEventListener("change", event => {{
      const file = event.target.files[0];
      if (!file) return;
      const reader = new FileReader();
      reader.onload = () => {{
        try {{
          state = normalizeState(JSON.parse(reader.result));
          saveState();
          renderAll();
        }} catch (error) {{
          alert("导入失败，文件不是可用的 JSON 数据。");
        }}
      }};
      reader.readAsText(file);
    }});
    document.getElementById("growthInput").addEventListener("change", event => {{
      state.targetGrowth = Number(event.target.value || 0) / 100;
      redistributeTargets();
    }});
    document.getElementById("skuSearchInput").addEventListener("input", event => {{
      skuFilters.search = event.target.value;
      if (activeView === "sku") {{
        renderSkuFilterSummary();
        renderParetoChart("skuChart");
        renderSkuTable();
        renderSkuGapTable();
      }}
    }});
    document.getElementById("skuSortSelect").addEventListener("change", event => {{
      skuFilters.sort = event.target.value;
      if (activeView === "sku") {{
        renderSkuFilterSummary();
        renderSkuTable();
        renderSkuGapTable();
      }}
    }});
    document.getElementById("skuOnlyGapInput").addEventListener("change", event => {{
      skuFilters.onlyGap = event.target.checked;
      if (activeView === "sku") {{
        renderSkuFilterSummary();
        renderParetoChart("skuChart");
        renderSkuTable();
        renderSkuGapTable();
      }}
    }});

    window.addEventListener("resize", () => {{
      renderMonthlyChart();
      if (activeView === "sku") renderParetoChart("skuChart");
    }});
    window.addEventListener("beforeunload", () => {{
      try {{ localStorage.setItem(STORAGE_KEY, JSON.stringify(state)); }} catch (error) {{}}
    }});

    renderAll();
  </script>
</body>
</html>"""


def main():
    workbook_path = Path(os.environ["XLSX_PATH"])
    state = build_state(workbook_path)
    output_path = workbook_path.parent / "GMA_2026_销售预测看板.html"
    output_path.write_text(html_template(state), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
